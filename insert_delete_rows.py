#!/usr/local/bin/python3

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim_insert.xlsx')
ws = wb.active

# Insert rows
ws.insert_rows(1)

# Delete rows
ws.delete_rows(6)


wb.save('tim_insert.xlsx')

