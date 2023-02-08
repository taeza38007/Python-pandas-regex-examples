#!/usr/local/bin/python3

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('./tim_example.xlsx')
ws = wb.active 


# Move from selected range of cells to new location
ws.move_range('A1:B5', rows=4, cols=2)
ws.delete_cols(1,2)

wb.save('tim_move.xlsx')