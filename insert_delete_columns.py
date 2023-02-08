#!/usr/local/bin/python3

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('./tim_column_insert.xlsx')
ws = wb.active


# Insert columns
ws.insert_cols(1,2)

# Delete columns
ws.delete_cols(3)



wb.save('tim_column_insert.xlsx')


