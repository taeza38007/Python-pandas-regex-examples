#!/usr/local/bin/python3
from openpyxl import Workbook, load_workbook

# Function to get column name
from openpyxl.utils import get_column_letter

# Load workbook file
wb = load_workbook('./grade.xlsx')
ws = wb.active

# Loop through worksheet 
for row in range(1, 6):
    for col in range(1, 5):
        # Assign column letter to column loop
        char = get_column_letter(col)
        # Print value in side each cells 
        print(ws[char + str(row)].value)


# wb.save('tim_example.xlsx')
