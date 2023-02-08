#!/usr/local/bin/python3  
from openpyxl import Workbook, load_workbook

# Load workbook xlsx file
workbook = load_workbook('./grade.xlsx')
# Set worksheet to use
#worksheet = workbook['sheet1']
worksheet = workbook.active 

# Print value in cell 'A1'
print(worksheet['A1'].value)

# Set new value to cell 'A1'
worksheet['A1'].value = 'Nick name'
print(worksheet['A1'].value)

# Create new worksheet
workbook.create_sheet('new sheet')

# Print all worksheet names
print(workbook.sheetnames)

# Save to xlsx file
workbook.save('grade.xlsx')