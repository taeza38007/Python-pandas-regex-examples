#!/usr/local/bin/python3

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# Import fonts
from openpyxl.styles import Font

# Dict of data
data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = 'Grades'

# Assign heading values from keys in dict
headings = ['Name'] + list(data['Bill'].keys())
# Append data to worksheet
ws.append(headings)
print(headings)

# Loop through each person in dict
for person in data:
    # Assign values of data for each person to 'grades' and convert to lists
    grades = list(data[person].values())
    # Print person's name
    print([person])
    # Print person's keys & values
    print(data[person])
    print(grades)

    # Append person'name + grades(values of each person's grade)
    ws.append([person] + grades)

# Loop through all column dynamically starting from column 'B' til the last one
for cols in range(2, len(data['Joe']) + 2):
    print(cols)
    # Assign column letters with numbers
    char = get_column_letter(cols)
    print(char)

    # Create new row for average for each person's values
    ws[char + str(len(data) + 2)] = f"=AVERAGE({char + str(len(data) - 3)}:{char + str(len(data) + 1)})"

# Loop through all columns
for col in range(1, len(data["Bill"]) + 2):
    # Apply style to each column's headings (Bold, text color Blue)
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color='000066BB')
    print(col)

# print(len(data['Joe']))
# print(data['Joe'])
# # print(data.keys())
# person_name = list(data.keys())
# print(person_name[0])
# print(len(data))


# Save to a file called "newGrades.xlsx"
wb.save('newGrades.xlsx')
