#!/usr/local/bin/python3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('./tim_example.xlsx')
ws = wb.active

# Merge cells
ws.merge_cells('A1:D1')
ws.merge_cells('A2:C2')
ws.merge_cells('C3:D3')

# Merge 2 rows 
ws.merge_cells('B4:D5')


# Unmerge cells

ws.unmerge_cells('A1:D1')
ws.unmerge_cells('C3:D3')



# Save file to 'tim_merge' 
wb.save('tim_merge.xlsx')
